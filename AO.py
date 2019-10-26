import openpyxl
import csv
import sqlite3
from datetime import datetime
import calendar
from PyQt5.QtWidgets import QApplication, QWidget, QHeaderView
from PyQt5 import uic, QtWidgets, QtCore
import os
import time
import subprocess
import glob


class FileTurni:
    """gestisce le operazioni dal file turni, prende in input il percorso del file turni .xlsx"""
    def __init__(self, tabellone):
        self.tabellone = tabellone

    def turni_mensili(self, dipendente):
        """
        elenca i turni del dipendente preso in input
        :param dipendente: stringa di tipo NOME COGNOME facendo attenzione alle maiuscole (es. Mario Rossi)
        :return: ritorna un dizionario contenente i turni di lavoro come data:turno (es. {'2019-09-02': '16:00-22:00'} )
        """

        fileturni = openpyxl.load_workbook(self.tabellone)       # file excel dei turni - es. Tabellone.xlsx
        foglio = fileturni.active                                # individua il foglio principale
        dizturni = {}                                            # dizionario turni
        for riga in foglio.iter_rows():                          # restituisce una tupla per ogni riga del tabellone
            for cella in riga:                                   # spacchetta le celle nella riga
                if cella.value == dipendente:                    # confronta ogni cella con variabile "dipendente"
                    for cella2 in riga:                          # per ogni cella nella riga trovata...
                        if "-" and ":" in str(cella2.value):     # cerca il turno nella riga del dipendente
                            colonna_turno = cella2.column        # individua la colonna della cella attuale(per la data)
                            data_turno = foglio.cell(column=colonna_turno, row=1)  # aggiunge la data (dalla prima riga)
                            dizturni[str(data_turno.value.date())] = cella2.value  # crea un dizionario con {data:turno}
        return dizturni

    def _lista_elementi_in_colonna(self, tipo):                      # metodo ad uso interno alla classe
        """
        elenca gli elementi in una colonna (non in riga come per il metodo .turni_mensili)
        :param tipo: Altamente raccomandato scegliere tra Nominativo, Contratto, Modulo e Skill
        :return: ritorna una lista degli elementi in colonna
        """
        tipo = tipo.capitalize()                                    # controlla le maiuscole
        fileturni = openpyxl.load_workbook(self.tabellone)          # file excel dei turni - es. Tabellone.xlsx
        foglio = fileturni.active                                   # individua il foglio principale
        listaelem = []
        for colonna in foglio.iter_cols():                          # restituisce una tupla per ogni riga del tabellone
            for cella in colonna:                                   # spacchetta le celle nella riga
                if cella.value == tipo:                             # confronta ogni cella con variabile "tipo"
                    for cella2 in colonna:
                        if "None" not in str(cella2.value):          # filtra i valori nulli (None)
                            if tipo not in str(cella2.value):        # evita di aggiungere il tipo di valore cercato
                                listaelem.append(cella2.value)       # aggiunge alla lista
        return listaelem

    def lista_elementi_in_tabellone(self, dipendente):
        """crea una lista con data e turno (es. '2019-10-07, 08:30-14:30',...)"""
        lista_elem = []
        for data, turno in self.turni_mensili(dipendente).items():
            lista_elem.append(f"{data}, {turno}")
        return lista_elem

    def elenco_dipendenti(self):
        """elenca i dipendenti"""
        return self._lista_elementi_in_colonna("Nominativo")

    def cerca_turno(self, dipendente, data):
        """cerca un turno attraverso nome dipendente e data"""
        dizturni = self.turni_mensili(dipendente)
        return {data: dizturni[data]}

    @staticmethod
    def verifica_parcheggio(data_da_verificare, turno_da_verificare):
        """
        Avvisa della mancanza di parcheggio per pulizie stradali.
        (trigger: secondo lunedi del mese con inizio turno entro le 11)
        :param data_da_verificare: Data del turno in formato YYYY-MM-DD (es.'2019-07-24')
        :param turno_da_verificare: inserire in formato "00:00-00:00" (es. '16:00-22:00')
        :return: restituisce,in base alla condizione trigger una stringa: " ! No parcheggio" oppure ""
        """
        data_input = datetime.strptime(data_da_verificare, "%Y-%m-%d")
        c1 = calendar.Calendar(firstweekday=calendar.MONDAY)
        lista_date_mensile = c1.monthdatescalendar(data_input.year, data_input.month)

        lista_lunedi = []
        for settimana in lista_date_mensile:  # lista dei giorni delle settimane incluse nel mese (una tupla per sett.)
            for data in settimana:                            # scompatta le tuple precedenti: crea un elenco di giorni
                if data.weekday() == calendar.MONDAY:         # ...se la data indicata è un LUNEDI
                    if data.month == data_input.month:        # ed è nel mese indicato...
                        lista_lunedi.append(data)             # aggiunge alla lista dei lunedi del (solo) mese corrente
        seclun = lista_lunedi[1]                              # 2' elemento della lista dei lunedi ossia il 2' del mese

        seclunbool = str(seclun) == str(data_da_verificare)            # è il 2' lunedi del mese? (True/False)
        turnchkbool = str(turno_da_verificare[:2]) <= str(11)          # il turno è entro le 11:00? (True/False)

        if seclunbool is True and turnchkbool is True:                 # se sono vere entrambe le condizioni
            return str(" ! No parcheggio")                             # stringa in caso non ci sia parcheggio
        else:
            return str("")                                             # stringa in caso ci sia parcheggio


class Tabella:
    """gestisce le operazioni dalla tabella dei turni, prende in input il percorso della tabella"""
    def __init__(self, tabella):
        self.tabella = tabella

    def elenca_righe(self):
        """ritorna una lista contenente un'altra lista con [turno, note, notifica] per ogni riga in tabella """
        with open(self.tabella) as filetabella:                 # legge il file csv
            lettorecsv = csv.reader(filetabella)
            next(lettorecsv)                                    # evita lettura della riga dell'indice
            dati_in_tabella = []
            for riga in lettorecsv:                             # crea una lista per ogni riga e aggiunge alla lista
                dati_in_tabella.append(riga)
            return dati_in_tabella

    def verifica_presenza_turno_su_tabella(self, turno):
        """
        verifica se presente il turno nella tabella e i suoi stati(es.sveglia,abilitazione)
        (serve in caso di nuovi turni/non presenti in tabella)
        :param turno: Inserire un turno in formato "00:00-00:00" (es. '07:00-13:00')
        :return: restituisce True se lo trova altrimenti False per avvisare che non è in lista
        """
        try:
            if turno == self.cerca_nella_tabella(turno)[0][0]:  # Controlla se presente in tabella
                return True
            else:
                return False
        except Exception as info_errore:
            print(info_errore)
            return False

    def verifica_sveglia(self, turno):
        """verifica se è attiva la sveglia (DEFAULT ATTIVA)"""
        try:
            if turno == self.cerca_nella_tabella(turno)[0][0] and self.cerca_nella_tabella(turno)[0][3] == "NO":
                return False
            else:
                return True
        except Exception as info_errore:
            print(info_errore)
            return True

    def cerca_nella_tabella(self, turno):
        """
        cerca nella il turno nella tabella
        :param turno: Inserire un turno in formato "00:00-00:00" (es. '07:00-13:00')
        :return: restituisce le informazioni su di esso: [turno, note, e ora notifica]
        """
        for elem in self.elenca_righe():
            if elem[0] == turno:
                return [elem]


class DBTurni:
    """gestisce le operazioni del database, prende in input il percorso del database"""
    def __init__(self, database):
        self.database = database

    def comando_sql(self, comando):
        """metodo interno per eseguire comandi sql custom, ritorna la risposta dal db."""
        ttdb = sqlite3.connect(self.database)         # si collega al database SQL
        ttdb_cursor = ttdb.cursor()
        ttdb_cursor.execute(comando)                  # esegue il comando
        risposta_ttdb = ttdb_cursor.fetchall()        # segna le risposte dal db
        ttdb.commit()                                 # scrive i dati,eventualmente
        ttdb.close()                                  # chiude il file
        return risposta_ttdb                          # ritorna la risposta del db

    def ottimizza_db(self):
        """ottimizza il database: elimina i vecchi turni inattivi"""
        return self.comando_sql("DELETE FROM reminders WHERE reminder_active='0';")

    def scrivi_turno(self, data, note, ora_notifica, parcheggio, perc_suoneria, sveglia):
        """
        Aggiunge un turno di lavoro al database
        :param data: Data del turno in formato YYYY-MM-DD (es.'2019-07-24')
        :param note: Note sul turno
        :param ora_notifica: orario di notifica in formato HH:MM (es. '15:25')
        :param parcheggio: Aggiunge nota su disponibilità parcheggio (es.' ! No parcheggio')
        :param perc_suoneria: Aggiunge il percorso su memoria disp. android della suoneria della notifica
        :param sveglia: Tipo booleano,si intende sveglia attiva? True/False
        :return: restituisce la risposta o i dati dal database
        """
        if sveglia is True:
            f = self.comando_sql(f"INSERT INTO reminders VALUES(NULL,'{note} {parcheggio}','{data} {ora_notifica}'"
                                 f",'1','0','0','0','0','11','','0','1','0','0','0','0','0','','0','1',"
                                 f"'{perc_suoneria}','1','5','1','1','0');")  # con sveglia
        else:
            f = self.comando_sql(f"INSERT INTO reminders VALUES(NULL,'{note} {parcheggio}','{data} {ora_notifica}'"
                                 f",'1','0','0','0','12','11','','0','1','0','0','0','0','0','','0','0',"
                                 f"'{perc_suoneria}','0','5','1','0','0');")  # no sveglia

        return f

    def _leggi_date_su_db(self):
        """ legge le date dal database e restituisce una lista """
        lista_reminder_date = self.comando_sql("SELECT reminder_date FROM reminders")
        lista_date = []
        for elem in lista_reminder_date:
            data = elem[0]                                      # ottiene il primo elemento dalla lista_reminder_date
            data_format = datetime.strptime(data, "%Y-%m-%d %H:%M").strftime("%Y-%m-%d")
            lista_date.append(data_format)
        return lista_date

    def lista_elementi_su_db(self):
        """ legge le date dal database e restituisce una lista """
        elem_su_db = self.comando_sql("SELECT reminder_date,reminder_name FROM reminders")
        lista_elem = []
        for elem, elem2 in elem_su_db:
            lista_elem.append(elem + ", " + elem2)
        return lista_elem

    def verifica_presenza_turno_su_db(self, data):                                    # (ex filtrodoppiturni)
        """
        verifica per ogni data fornita se il turno è presente nel db
        utile per evitare di duplicare i turni del mese se sono già presenti (anche in parte) nel nuovo Tabellone

        :param data: Data del turno in formato YYYY-MM-DD (es.'2019-07-24')
        :return: se presente nel database restituisce True,altrimenti False
        """
        if data in self._leggi_date_su_db():
            return True
        else:
            return False

    def cerca_turno(self, data):                                                              # Funzionalita aggiuntiva
        """
        legge uno o piu turni dal database
        :param data: Data del turno in formato YYYY-MM-DD (es.'2019-07-24')
        :return: restituisce una lista con dentro una o piu tuple per la data richiesta(normalmente una tupla)
        """
        f = self.comando_sql(f'SELECT reminder_date,reminder_name FROM reminders WHERE reminder_date LIKE "%{data}%";')
        return f

    @staticmethod
    def cerca_filedb_da_rimuovere(perc_filedb):
        """
        verifica se ci sono vecchi file database nella stessa directory del file indicato
        :param perc_filedb: percorso del file database
        :return: Ritorna una lista vuota se non ci sono file vecchi altrimenti una lista di essi
        """
        # ELENCA TUTTI I FILE NELLA DIRECTORY DEL FILE SELEZIONATO
        directory = os.path.dirname(perc_filedb)                # seleziona la directory del file selezionato
        listafiledir = glob.glob(directory + "\\*")             # elenca in una lista i file presenti nella directory

        # ELENCA SOLO I FILE (no cartelle) CONTENENTI "TimeTune Backup" E CHE NON HANNO ESTENSIONE es(.txt)
        filestt = []
        for filex in listafiledir:
            if os.path.isfile(filex):    # controlla che siano file e non cartelle
                if "TimeTune Backup" in filex:  # controlla che ci sia "TimeTune Backup"
                    if not os.path.splitext(filex)[1]:      # controlla che il file non abbia una estensione
                        filestt.append(filex)

        # CERCA IL FILE PIU RECENTE E LO ESCLUDE.. ELENCA TUTTI GLI ALTRI FILE MENO RECENTI
        file_vecchi = []
        file_nuovo = max(filestt, key=os.path.getctime)  # restituisce il piu recente
        for file in filestt:                             # rimuove i vecchi file
            if file != file_nuovo:                       # tranne quello piu recente
                file_vecchi.append(file)

        if file_vecchi:
            return file_vecchi
        else:
            return []

    @staticmethod
    def cerca_filedb_piu_recente(directory):
        """cerca il file piu recente nella directory fornita"""
        # ELENCA TUTTI I FILE NELLA DIRECTORY DEL FILE SELEZIONATO
        listafiledir = glob.glob(directory + "//*")             # elenca in una lista i file presenti nella directory

        # ELENCA SOLO I FILE (no cartelle) CONTENENTI "TimeTune Backup" E CHE NON HANNO ESTENSIONE es(.txt)
        filestt = []
        for filex in listafiledir:
            if os.path.isfile(filex):  # controlla che siano file e non cartelle
                if "TimeTune Backup" in filex:  # controlla che ci sia "TimeTune Backup"
                    if not os.path.splitext(filex)[1]:      # controlla che il file non abbia una estensione
                        filestt.append(filex)

        # CERCA IL FILE PIU RECENTE
        file_nuovo = max(filestt, key=os.path.getctime)  # restituisce il piu recente
        return file_nuovo


class ManagerTurni:
    """
    gestisce le operazioni sui turni (li inserisce su db,esegue i cambi turno e altre funzioni di alto livello)
    :param dipendente: stringa con COGNOME e NOME del dipendente (es. mario rossi)
    :param fileturni: istanza di FileTurni
    :param filetabella: istanza di FileTurni
    :param dbturnimensile: istanza di DBTurni
    :param perc_suoneria: percorso del file suoneria per la notifica
    """
    def __init__(self, dipendente, fileturni, filetabella, dbturnimensile, perc_suoneria):
        self.dipendente = str(dipendente)
        self.fileturni = fileturni
        self.filetabella = filetabella
        self.dbturnimensile = dbturnimensile
        self.perc_suoneria = perc_suoneria
        window.aggiorna_lineedit_suoneria()

    def inserisci_tutti_i_turni_su_db(self):
        """
        inserisce tutti i turni sul database
        :return ritorna piu liste con l'esito delle operazioni: dizturni, turni_scritti, turni_saltati, errori
        I turni_saltati sono quelli già inclusi nel database mentre gli errori sono quelli non inclusi nella tabella.
        """

        dizturni = self.fileturni.turni_mensili(self.dipendente)                     # dizionario con i turni mensili

        errori = []
        turni_scritti = []
        turni_saltati = []

        for data, turno in dizturni.items():                                         # restituisce data e turno singoli
            if self.filetabella.verifica_presenza_turno_su_tabella(turno) is False:  # c.errori: in caso di nuovi turni
                errori.append(f"{data}, {turno}")

            elif self.dbturnimensile.verifica_presenza_turno_su_db(data) is False:   # controlla se la data è già nel db
                turno_da_scrivere = self.filetabella.cerca_nella_tabella(turno)
                note = turno_da_scrivere[0][1]
                notifica = turno_da_scrivere[0][2]
                parcheggio = self.fileturni.verifica_parcheggio(data, turno)

                sveglia = self.filetabella.verifica_sveglia(turno)  # controlla sveglia
                self.dbturnimensile.scrivi_turno(data, note, notifica, parcheggio, self.perc_suoneria, sveglia)
                turni_scritti.append(f"{data}, {turno}")                             # aggiunge i turni scritti
            else:
                turni_saltati.append(f"{data}, {turno}")                             # indica eventuali turni saltati

        lista_turni = []
        for data, turno in dizturni.items():
            lista_turni.append(f"{data}, {turno}")

        self.dbturnimensile.ottimizza_db()                                           # elimina vecchi turni sul database

        return lista_turni, turni_scritti, turni_saltati, errori


class Ui(QWidget):
    def __init__(self):
        super().__init__()
        try:
            uic.loadUi("AO_files\\AO.ui", self)
        except FileNotFoundError:
            print("Errore: File AO_files\\AO.ui non trovato")
            time.sleep(5)
        self.comboBox.activated[str].connect(self.cambio_nome_dip_combobox)  # collega le variazioni della combobox...
        self.ricarica_tabella()
        self.pushButton_2.setEnabled(False)
        self.pushButton_3.setEnabled(False)
        self.pushButton_4.setEnabled(False)
        self.pushButton_9.setEnabled(False)
        if not os.path.exists(f"{os.environ['HOMEDRIVE']}\\Program Files\\Google\\Drive\\googledrivesync.exe"):
            self.pushButton_drivesync.setText("Google Drive sync non disponibile")
            self.pushButton_drivesync.setEnabled(False)
        self._google_drive_run_check()  # controlla se avviato e imposta testo pulsante (setText nel metodo)
        self.lineEdit_suoneria.setReadOnly(True)
        self.aggiorna_lineedit_suoneria()

    filetabella1 = None  # dichiara valore default delle variabili globali (evita errore pep8 global)
    fileturni1 = None
    nome_dip2 = None
    perc_filedb_fixed = None
    filedb1 = None

    @staticmethod
    def modifica_suoneria_pulsante():
        try:
            os.startfile("AO_files\\config_suoneria.txt")
        except Exception as info_errore:
            print(info_errore)

    def aggiorna_lineedit_suoneria(self):
        with open("AO_files\\config_suoneria.txt") as suoneria:
            perc_suoneria = suoneria.read()
        self.lineEdit_suoneria.setText(perc_suoneria)

    def _google_drive_run_check(self):
        listatask = subprocess.check_output("tasklist")
        gdriveisrunning = "googledrivesync.exe" in str(listatask)
        if gdriveisrunning is True:
            self.pushButton_drivesync.setText("Google Drive sync già avviato!")
            self.pushButton_drivesync.setEnabled(False)
        return gdriveisrunning

    def ricarica_tabella(self):
        try:
            global filetabella1
            filetabella1 = Tabella("AO_files\\Tabella.csv")
            self.tableWidget.clear()
            for indice, elem in enumerate(filetabella1.elenca_righe()):  # imposta il numero di righe della tabella
                self.tableWidget.setRowCount(indice + 1)  # aggiunge una riga (mostra TUTTE le righe,inclusa ultima)
                self.tableWidget.setColumnCount(4)
                self.tableWidget.setItem(indice, 0, QtWidgets.QTableWidgetItem(elem[0]))
                self.tableWidget.setItem(indice, 1, QtWidgets.QTableWidgetItem(elem[1]))
                self.tableWidget.setItem(indice, 2, QtWidgets.QTableWidgetItem(elem[2]))
                self.tableWidget.setItem(indice, 3, QtWidgets.QTableWidgetItem(elem[3]))
            self.tableWidget.setHorizontalHeaderLabels(["TURNO", "NOTE", "NOTIFICA", "SVEGLIA"])
            self.tableWidget.resizeColumnsToContents()  # resize delle colonne tab.turni
            self.tableWidget.resizeRowsToContents()  # resize delle righe tab.turni
            self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # adatta la tab.alla finestra
        except FileNotFoundError:
            print("Errore: File AO_files\\Tabella.csv non trovato")
            time.sleep(5)

    def carica_tabellone(self):
        try:
            global fileturni1
            fileturni, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Seleziona file...', QtCore.QDir.rootPath(),
                                                                 "Excel files (*.xlsx);;ALL files (*.*)")
            fileturni1 = FileTurni(fileturni)
            self.comboBox.clear()
            self.comboBox.addItem("Seleziona...")
            self.comboBox.addItems(fileturni1.elenco_dipendenti())  # aggiunge i dipendenti alla lista
            self.pushButton_2.setText("Cambia...")

            self.pushButton_3.setEnabled(True)  # abilita tasto inserisci turni su db
        except Exception as info_errore:
            print(info_errore)
            QtWidgets.QMessageBox.warning(window, "Errore", "File vuoto o non riconosciuto!")

    def cambio_nome_dip_combobox(self, nome_dip):
        try:
            global nome_dip2
            nome_dip2 = str(nome_dip)
            self.listWidget.clear()
            self.listWidget.addItems(fileturni1.lista_elementi_in_tabellone(nome_dip))  # aggiunge i turni
        except Exception as info_errore:
            print(info_errore)
            QtWidgets.QMessageBox.warning(window, "Errore", 'Lista vuota (per iniziare clicca "Seleziona...")'
                                                            ' o dipendente non trovato!')

    def carica_database(self):

        if not self._google_drive_run_check():      # controlla che google drive non sia gia avviato
            QtWidgets.QMessageBox.information(window, "Info",
                                              "Google drive sync non avviato.\n\n"
                                              "Puoi avviarlo e utilizzarlo per "
                                              "sincronizzare il file di backup...")

        QtWidgets.QMessageBox.information(window, "Info", "Esegui il backup del database dall'app Timetune...\n\n"
                                                          "Premi OK per selezionare il database")
        try:
            global perc_filedb_fixed
            perc_cart_filedb = QtWidgets.QFileDialog.getExistingDirectory(window, 'Seleziona cartella...')
            perc_filedb = DBTurni.cerca_filedb_piu_recente(perc_cart_filedb)
            perc_filedb_fixed = perc_filedb.replace('\\', '/')

            global filedb1
            filedb1 = DBTurni(str(perc_filedb_fixed))
            self.listWidget_3.clear()
            self.listWidget_3.addItems(filedb1.lista_elementi_su_db())
            self.pushButton.setText("Cambia...")
            self.pushButton_4.setEnabled(True)  # abilita tasto aggiorna per db
            self.pushButton_2.setEnabled(True)  # abilita tasto per selezionare db
            self.pushButton_9.setEnabled(True)  # abilita tasto per selezionare comandi sql manuali
            self._google_drive_run_check()

            if window3.aggiorna_finestra():  # aggiorna la finestra e se ci sono file da rimuovere mostra window3
                window3.show()

        except Exception as info_errore:
            print(info_errore)
            QtWidgets.QMessageBox.warning(window, "Errore", "Nessuna cartella selezionata o file non riconosciuto!")

    def inserisci_turni_pulsante(self):
        try:
            perc_suoneria = self.lineEdit_suoneria.text()
            manager1 = ManagerTurni(nome_dip2, fileturni1, filetabella1, filedb1, perc_suoneria)
            lista_turni, turni_scritti, turni_saltati, errori = manager1.inserisci_tutti_i_turni_su_db()
            self.aggiorna_gui_turni()
            self.listWidget_2.clear()
            self.listWidget_2.addItem(f"Turni trovati: {len(lista_turni)}")
            self.listWidget_2.addItem(f"Turni scritti: {len(turni_scritti)}")
            self.listWidget_2.addItem(f"Turni saltati(già su db): {len(turni_saltati)}")
            self.listWidget_2.addItem(f"Turni non in lista(errori): {len(errori)}")
            self.listWidget_7.clear()
            self.listWidget_6.clear()
            self.listWidget_5.clear()
            self.listWidget_4.clear()
            self.listWidget_7.addItems(lista_turni)
            self.listWidget_6.addItems(turni_scritti)
            self.listWidget_5.addItems(turni_saltati)
            self.listWidget_4.addItems(errori)
            if len(errori) == 0 and len(turni_saltati) == 0 and len(turni_scritti) >= 1:
                QtWidgets.QMessageBox.information(window, "Info", "Operazione eseguita con successo.\n\n"
                                                          "Ripristina ora il backup sull'app Timetune!")
            elif len(errori) == 0 and len(turni_saltati) >= 1:
                QtWidgets.QMessageBox.information(window, "Info", "Operazione eseguita.\n\n"
                                                          "Alcuni turni sono stati saltati perchè le date erano "
                                                          "già presenti sul database...\n"
                                                          "Verifica i turni scritti prima di ripristinare il backup "
                                                          "sull'app Timetune ed eventualmente modificali!")
            else:
                QtWidgets.QMessageBox.warning(window, "Info", "Operazione eseguita con errori!\n\n"
                                                      "Verifica se ci sono ad esempio turni nuovi ed eventualmente "
                                                      "inseriscili in Tabella,quindi ripeti l'operazione...")

        except Exception as info_errore:
            print(info_errore)
            QtWidgets.QMessageBox.warning(window, "Errore", "Errore nella scrittura del database!")

    def aggiorna_gui_turni(self):
        self.listWidget_3.clear()
        self.listWidget_3.addItems(filedb1.lista_elementi_su_db())
        self.listWidget_3.sortItems()

    @staticmethod
    def modifica_tabella_pulsante():
        try:
            os.startfile("AO_files\\Tabella.csv")
        except FileNotFoundError:
            QtWidgets.QMessageBox.warning(window, "Errore", 'File "AO_files\\Tabella.csv" non trovato')

    @staticmethod
    def info_pulsante():
        QtWidgets.QMessageBox.information(window, "Info", """
        Descrizione:
        Il programma aggiunge dei turni di lavoro ad
        un database dell'app (di terze parti) Timetune per Android.
        I turni vengono letti da un file excel.

        Viene impostato un reminder per ogni turno trovato
        con notifica e note personalizzati (in base ai dati presenti
        nella Tabella.)

        Il sistema inoltre aggiunge automaticamente alle note l'eventuale 
        mancanza di parcheggio in caso di pulizie stradali
        (Secondo lunedì del mese con inizio turno entro le 11:00)
        
        Versione: 1.0
        Il programma viene fornito senza alcuna garanzia di funzionamento,
        è Open Source in licenza GNU GPL V3. Utilizza le librerie Qt e altro 
        software (di terze parti) TimeTune per Android e Google drive per windows.
        
        (c) 2019 Graziano Porcu
        https://github.com/Grawa
        Contatto: VGPLabs@gmail.com
        """)

    @staticmethod
    def guida_pulsante():
        QtWidgets.QMessageBox.information(window, "Info", """
        Guida:
        
        0) Importante per il primo utilizzo:
            Assicurarsi che il file della suoneria si trovi nel percorso corretto
            sul dispositivo,altrimenti non potrà suonare la sveglia.
            Eventualmente copiare o spostare il file della suoneria sul dispositivo...
            (di default il percorso è sdcard/Ringtones/suoneria.ogg) 
            
        1) Fare il backup del database da cellulare (consigliato su Google Drive*)...
            (app TimeTune>Impostazioni>Backup)
           
           
        2) Selezionare la cartella dove presente il database del backup*
            (se presenti più file verrà individuato automaticamente il più recente)
            e il file del Tabellone...
            Selezionare quindi il dipendente da inserire.
            Cliccare quindi su "Inserisci turni su database".
           
            Nota: è importante creare sempre il backup del database aggiornato 
            prima di scrivere i turni nuovi, in alternativa si perderanno 
            eventuali cambi e variazioni di turno eseguite tramite l'app.
            
           
        3) Ripristinare il database sul cellulare...
            (app TimeTune>Impostazioni>Backup)
            
            Nota: In caso di backup su Google Drive**, da pc accertarsi
            che il backup sia stato caricato correttamente da pc e sul
            cellulare di scaricare la copia aggiornata del file.
            
        ---------------------------
        
        Note:
        
        *Non rinominare il file di backup: per essere riconosciuto  
        automaticamente deve contenere la parola "Timetune Backup"

        **Integrazione con google drive sync disponibile per PC Windows.
          Link: https://www.google.com/drive/download/
            """)

    @staticmethod
    def googledrivesync_pulsante():
        os.startfile(f"{os.environ['HOMEDRIVE']}\\Program Files\\Google\\Drive\\googledrivesync.exe")
        QtWidgets.QMessageBox.information(window, "Info", "Avvio di Google Drive sync...\n\nNota:\n"
                                                  "Il programma potrebbe avviarsi ridotto a icona.")

    @staticmethod
    def comandi_sql_manuali_pulsante():
        window2.show()


class UiComandiSql(QWidget):
    def __init__(self):
        super().__init__()
        try:
            uic.loadUi("AO_files\\CSQL.ui", self)
        except FileNotFoundError:
            print("Errore: File AO_files\\CSQL.ui non trovato")
            time.sleep(5)

    def invio_pulsante(self):
        comando = self.lineEdit.text()
        try:
            risposta = filedb1.comando_sql(str(comando))
            if not any(risposta):
                self.textBrowser.setText("COMANDO SQL INVIATO (nessuna risposta dal database)")
            else:
                self.textBrowser.setText(str(risposta))
            window.aggiorna_gui_turni()
        except Exception as info_errore:
            print(info_errore)
            self.textBrowser.setText(f"COMANDO SQL NON RICONOSCIUTO.")

    def elimina_pulsante(self):
        self.lineEdit.setText("DELETE FROM reminders WHERE _id=' ' ")

    def eliminatutti_pulsante(self):
        self.lineEdit.setText("DELETE FROM reminders ")

    def seleziona_pulsante(self):
        self.lineEdit.setText("SELECT * FROM reminders")

    def cerca_pulsante(self):
        self.lineEdit.setText("SELECT * FROM reminders WHERE reminder_date LIKE '%YYYY-MM-DD%'")


class UiEliminaVecchiDB(QWidget):
    def __init__(self):
        super().__init__()
        self.lista_file_da_elim = None
        try:
            uic.loadUi("AO_files\\ELIM_FILE_DB.ui", self)
        except FileNotFoundError:
            print("Errore: AO_files\\ELIM_FILE_DB.ui non trovato")
            time.sleep(5)

    def aggiorna_finestra(self):
        """
        aggiorna la lista di file da eliminare
        :return: Controlla se ci sono file da eliminare ritorna True altrimenti False """

        self.lista_file_da_elim = DBTurni.cerca_filedb_da_rimuovere(perc_filedb_fixed)  # Verif.se ci sono file da elim.
        self.listWidget.clear()
        self.listWidget.addItems(self.lista_file_da_elim)  # Aggiunge la lista dei file al listWidget

        if self.lista_file_da_elim:  # ritorna True se ci sono file da eliminare altrimenti False
            return True
        else:
            return False

    def mantieni_files(self):  # pulsante NO
        self.aggiorna_finestra()
        self.close()

    def elimina_files(self):  # pulsante Elimina
        for file in self.lista_file_da_elim:
            os.remove(file)
            self.aggiorna_finestra()
            self.close()


if __name__ == "__main__":
    app = QApplication([])
    window = Ui()
    window2 = UiComandiSql()
    window3 = UiEliminaVecchiDB()
    window.show()
    app.exec()
